"""目标数据库工具：动态连接、Schema 提取、SQL 执行，并支持 Excel 作为可查询数据源。"""
import os
import re
from typing import List, Tuple

from sqlalchemy import create_engine, inspect, text
from sqlalchemy.pool import StaticPool

from .security import aes_decrypt


def _build_url(cfg: dict) -> str:
    t = cfg["type"]
    pwd = cfg.get("password") or ""
    if t == "mysql":
        port = cfg.get("port") or 3306
        return f"mysql+pymysql://{cfg.get('username')}:{pwd}@{cfg.get('host')}:{port}/{cfg.get('database')}"
    if t == "postgresql":
        port = cfg.get("port") or 5432
        return f"postgresql+psycopg2://{cfg.get('username')}:{pwd}@{cfg.get('host')}:{port}/{cfg.get('database')}"
    if t == "sqlite":
        return f"sqlite:///{cfg.get('file_path') or cfg.get('database')}"
    raise ValueError(f"不支持的数据库类型: {t}")


def _excel_to_sqlite(file_path: str):
    """把 Excel 每个工作表读入内存 SQLite，每个表对应一个 sheet。返回 engine。"""
    from openpyxl import load_workbook

    if not os.path.exists(file_path):
        raise FileNotFoundError("文件不存在")

    # StaticPool + 单连接：保证内存 SQLite 在整个 engine 生命周期内持久（否则每次取连接都是空库）
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    wb = load_workbook(file_path, read_only=True, data_only=True)
    with engine.begin() as conn:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
            table = _safe_ident(ws.title)
            cols = ", ".join(f'"{_safe_ident(h)}" TEXT' for h in headers)
            conn.execute(text(f'CREATE TABLE "{table}" ({cols})'))
            placeholders = ", ".join(f":c{i}" for i in range(len(headers)))
            insert_sql = text(f'INSERT INTO "{table}" VALUES ({placeholders})')
            for row in rows[1:]:
                params = {f"c{i}": (row[i] if i < len(row) else None) for i in range(len(headers))}
                conn.execute(insert_sql, params)
    wb.close()
    return engine


def _safe_ident(name: str) -> str:
    return re.sub(r"[^\w一-鿿]", "_", str(name)).strip("_") or "col"


def build_engine(cfg: dict):
    """根据配置（password 字段为 AES 密文）创建目标数据库 engine。"""
    if cfg["type"] == "excel":
        return _excel_to_sqlite(cfg.get("file_path"))
    decrypted = dict(cfg)
    decrypted["password"] = aes_decrypt(cfg.get("password") or "")
    return create_engine(_build_url(decrypted), pool_pre_ping=True)


def get_schema(engine) -> str:
    """提取库内所有表结构，生成给 LLM 用的文本描述。"""
    inspector = inspect(engine)
    lines: List[str] = []
    for table in inspector.get_table_names():
        cols = inspector.get_columns(table)
        col_desc = ", ".join(f"{c['name']} {str(c['type'])}" for c in cols)
        lines.append(f"表 {table}({col_desc})")
    return "\n".join(lines) if lines else "（数据库中没有任何数据表）"


def run_query(engine, sql: str, limit: int = 200) -> Tuple[List[str], List[list]]:
    """执行只读 SQL，返回 (列名, 数据行)。仅允许 SELECT。"""
    cleaned = sql.strip().rstrip(";")
    if not cleaned.lower().startswith(("select", "with")):
        raise ValueError("仅支持查询(SELECT)语句")
    with engine.connect() as conn:
        result = conn.execute(text(cleaned))
        columns = list(result.keys())
        rows = [list(r) for r in result.fetchmany(limit)]
    return columns, rows

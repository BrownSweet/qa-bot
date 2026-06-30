"""数据导入：解析 Excel / 预览工作表数据。"""
import os

from fastapi import APIRouter, Depends

from .. import schemas, security
from ..utils import api_error

router = APIRouter(prefix="/api/excel", tags=["excel"])


def _detect_type(values) -> str:
    sample = [v for v in values if v is not None][:20]
    if not sample:
        return "string"
    if all(isinstance(v, (int, float)) for v in sample):
        return "number"
    from datetime import date, datetime
    if all(isinstance(v, (date, datetime)) for v in sample):
        return "date"
    return "string"


@router.post("/parse")
def parse_excel(body: schemas.ParseExcelRequest, user=Depends(security.get_current_user)):
    if not os.path.exists(body.file_path):
        raise api_error(404, "not_found", "文件不存在")
    from openpyxl import load_workbook
    try:
        wb = load_workbook(body.file_path, read_only=True, data_only=True)
    except Exception as e:
        raise api_error(400, "bad_request", f"无法解析文件：{e}")
    sheets = []
    for ws in wb.worksheets:
        sheets.append({"name": ws.title, "row_count": ws.max_row or 0, "column_count": ws.max_column or 0})
    wb.close()
    return {"sheets": sheets}


@router.post("/sheet-data")
def sheet_data(body: schemas.GetSheetDataRequest, user=Depends(security.get_current_user)):
    if not os.path.exists(body.file_path):
        raise api_error(404, "not_found", "文件不存在")
    from openpyxl import load_workbook
    wb = load_workbook(body.file_path, read_only=True, data_only=True)
    if body.sheet_name not in wb.sheetnames:
        wb.close()
        raise api_error(404, "not_found", "工作表不存在")
    ws = wb[body.sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {"columns": [], "rows": [], "total_rows": 0}

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:]
    limit = body.limit or 100
    preview = [list(r) for r in data_rows[:limit]]
    columns = [
        {"name": headers[i], "index": i,
         "type": _detect_type([r[i] if i < len(r) else None for r in data_rows])}
        for i in range(len(headers))
    ]
    return {"columns": columns, "rows": preview, "total_rows": len(data_rows)}
